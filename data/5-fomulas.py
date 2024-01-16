from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 1 - Lê pasta de trabalho e planilha
wb = load_workbook(filename='C:/Users/User/Desktop/python_estudos/minicurso_Onebitecode/data/barchart.xlsx')
sheet = wb['Relatorio']

# 2 - Referências das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
print(min_column)
print(max_column)
min_row = wb.active.min_row
max_row = wb.active.max_row
print(min_row, max_row)

# 3 - Incluíndo fórmula
for i in range(min_column+1, max_column+1):
  letter = get_column_letter(i)
  sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
  sheet[f"{letter}{max_row+1}"].style = "Currency"

  wb.save("C:/Users/User/Desktop/python_estudos/minicurso_Onebitecode/data/teste.xlsx")
