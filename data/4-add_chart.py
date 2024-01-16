from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1 - Lê pasta de trabalho e planilha
wb = load_workbook(filename='C:/Users/User/Desktop/python_estudos/minicurso_Onebitecode/data/pivot_table.xlsx')
sheet = wb['Relatorio']

# 2 - Referências das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
print(min_column)
print(max_column)
min_row = wb.active.min_row
max_row = wb.active.max_row
print(min_row, max_row)

# 3 - Adicionando Dados e Categorias no Gráfico
BarChart = BarChart()

data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)

categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

BarChart.add_data(data, titles_from_data=True)
BarChart.set_categories(categories)

# 4 - Exportando gráfico em arquivo excel
sheet.add_chart(BarChart, "B10")
BarChart.title = "Vendas por Fabricante"
BarChart.style = 2


# 5 - Salvando arquivo
wb.save("C:/Users/User/Desktop/python_estudos/minicurso_Onebitecode/data/barchart.xlsx")

